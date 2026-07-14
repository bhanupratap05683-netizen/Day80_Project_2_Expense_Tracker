

## 7. Final Python File: `day80_build.py`

"""
Day 80: Portfolio Project 2 - Expense Tracker (Build)
Scope: Categorization + Reporting
Focus: Regex rules, budget variance, multi-sheet Excel, embedded charts
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for stability
import matplotlib.pyplot as plt
import re
import os
import logging
from datetime import datetime, timedelta
import random

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage

# =============================================================================
# CONFIGURATION
# =============================================================================
CATEGORY_RULES = {
    'Groceries': ['supermarket', 'grocery', 'mart', 'food bazaar', 'whole foods'],
    'Dining': ['restaurant', 'cafe', 'coffee', 'pizza', 'burger', 'starbucks'],
    'Transport': ['uber', 'lyft', 'taxi', 'fuel', 'gas', 'metro', 'bus'],
    'Utilities': ['electric', 'water', 'internet', 'phone', 'gas bill'],
    'Entertainment': ['netflix', 'spotify', 'cinema', 'movie', 'game'],
    'Healthcare': ['pharmacy', 'doctor', 'hospital', 'medical', 'dental'],
    'Shopping': ['amazon', 'flipkart', 'mall', 'retail', 'clothing'],
    'Rent': ['rent', 'lease', 'housing'],
}

BUDGET_LIMITS = {
    'Groceries': 500,
    'Dining': 300,
    'Transport': 200,
    'Utilities': 150,
    'Entertainment': 100,
    'Healthcare': 200,
    'Shopping': 400,
    'Rent': 1200,
    'Miscellaneous': 100
}

INPUT_FILE = 'day80_input.xlsx'
OUTPUT_FILE = 'day80_expense_report.xlsx'
CHART_DIR = 'day80_charts'

# =============================================================================
# LOGGING SETUP
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('expense_tracker_day80.log', mode='w'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# STEP 1: PRACTICE SHEET GENERATION
# =============================================================================
def create_practice_sheet(output_path=INPUT_FILE):
    """Generate realistic raw expense data and budget sheet."""
    logger.info("Generating practice sheet...")
    random.seed(42)
    np.random.seed(42)

    descriptions_pool = [
        'Whole Foods Market', 'Starbucks Coffee', 'Uber Ride',
        'Electric Company', 'Netflix Subscription', 'CVS Pharmacy',
        'Amazon Purchase', 'Monthly Rent', 'Gas Station',
        'Local Restaurant', 'Grocery Store', 'Movie Theater',
        'Doctor Visit', 'Phone Bill', 'Metro Card', 'Unknown Vendor'
    ]

    data = []
    start_date = datetime(2026, 4, 1)
    for i in range(100):
        desc = random.choice(descriptions_pool)
        txn_date = start_date + timedelta(days=random.randint(0, 59))
        amount = round(random.uniform(5, 1200), 2)
        payment = random.choice(['Credit Card', 'Debit Card', 'Cash', 'UPI'])
        data.append({
            'Transaction_ID': f'TXN_{i+1:04d}',
            'Date': txn_date.strftime('%Y-%m-%d'),
            'Description': desc,
            'Amount': amount,
            'Payment_Method': payment
        })

    df_expenses = pd.DataFrame(data)
    df_budget = pd.DataFrame([
        {'Category': cat, 'Budget_Limit': limit}
        for cat, limit in BUDGET_LIMITS.items()
    ])

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_expenses.to_excel(writer, sheet_name='Raw_Expenses', index=False)
        df_budget.to_excel(writer, sheet_name='Budget', index=False)

    logger.info(f"Practice sheet created: {output_path}")


# =============================================================================
# STEP 2: DATA LOADING & VALIDATION
# =============================================================================
def load_data(filepath):
    """Load and validate input Excel workbook."""
    logger.info(f"Loading data from {filepath}")
    try:
        df_expenses = pd.read_excel(filepath, sheet_name='Raw_Expenses')
        df_budget = pd.read_excel(filepath, sheet_name='Budget')

        # Validation
        required_cols = {'Description', 'Amount', 'Date'}
        missing = required_cols - set(df_expenses.columns)
        if missing:
            raise ValueError(f"Missing required columns in Raw_Expenses: {missing}")

        df_expenses['Date'] = pd.to_datetime(df_expenses['Date'], errors='coerce')
        df_expenses['Amount'] = pd.to_numeric(df_expenses['Amount'], errors='coerce')

        if df_expenses['Date'].isna().any() or df_expenses['Amount'].isna().any():
            raise ValueError("Data type conversion failed: check Date and Amount columns")

        logger.info(f"Loaded {len(df_expenses)} transactions and {len(df_budget)} budget categories")
        return df_expenses, df_budget

    except Exception as e:
        logger.error(f"Data loading failed: {e}")
        raise


# =============================================================================
# STEP 3: CATEGORIZATION ENGINE
# =============================================================================
def categorize_expense(description, rules):
    """Apply regex-based keyword matching to categorize a transaction."""
    desc_lower = str(description).lower()
    for category, keywords in rules.items():
        for keyword in keywords:
            pattern = r'\b' + re.escape(keyword.lower()) + r'\b'
            if re.search(pattern, desc_lower):
                return category
    return 'Miscellaneous'


# =============================================================================
# STEP 4: PROCESSING & VARIANCE ANALYSIS
# =============================================================================
def process_expenses(df_expenses, df_budget, rules):
    """Categorize expenses, merge budgets, and compute variances."""
    logger.info("Processing expenses and calculating variances...")

    # Categorize
    df_expenses = df_expenses.copy()
    df_expenses['Category'] = df_expenses['Description'].apply(
        lambda x: categorize_expense(x, rules)
    )

    # Aggregate actual spend per category
    category_spend = df_expenses.groupby('Category')['Amount'].sum().reset_index()
    category_spend.columns = ['Category', 'Actual_Spend']

    # Merge with budget
    df_summary = df_budget.merge(category_spend, on='Category', how='outer')
    df_summary['Budget_Limit'] = df_summary['Budget_Limit'].fillna(0)
    df_summary['Actual_Spend'] = df_summary['Actual_Spend'].fillna(0)

    # Variance calculations
    df_summary['Variance'] = df_summary['Actual_Spend'] - df_summary['Budget_Limit']
    df_summary['Variance_Pct'] = np.where(
        df_summary['Budget_Limit'] == 0,
        np.nan,
        (df_summary['Variance'] / df_summary['Budget_Limit']) * 100
    )
    df_summary['Status'] = np.where(
        df_summary['Variance'] > 0, 'Over Budget',
        np.where(df_summary['Variance'] < 0, 'Under Budget', 'On Target')
    )

    # Merge budget limit back to processed transactions for audit trail
    df_processed = df_expenses.merge(
        df_budget[['Category', 'Budget_Limit']], on='Category', how='left'
    )
    df_processed['Budget_Limit'] = df_processed['Budget_Limit'].fillna(0)

    logger.info("Processing complete")
    return df_processed, df_summary


# =============================================================================
# STEP 5: AUTOMATED INSIGHTS
# =============================================================================
def generate_insights(df_summary):
    """Generate high-level financial insights."""
    total_spend = df_summary['Actual_Spend'].sum()
    total_budget = df_summary['Budget_Limit'].sum()
    insights = [
        f"Total Spend: ${total_spend:,.2f} | Total Budget: ${total_budget:,.2f}",
        f"Net Position: ${total_spend - total_budget:,.2f}"
    ]

    over_budget = df_summary[df_summary['Status'] == 'Over Budget']
    if not over_budget.empty:
        worst = over_budget.loc[over_budget['Variance'].idxmax()]
        insights.append(
            f"Highest Overrun: {worst['Category']} (${worst['Variance']:,.2f} over budget)"
        )

    for insight in insights:
        logger.info(f"INSIGHT: {insight}")
    return insights


# =============================================================================
# STEP 6: VISUALIZATION GENERATION
# =============================================================================
def generate_visualizations(df_summary, output_dir=CHART_DIR):
    """Create budget vs actual and variance charts."""
    logger.info("Generating visualizations...")
    os.makedirs(output_dir, exist_ok=True)

    # Chart 1: Budget vs Actual
    fig, ax = plt.subplots(figsize=(10, 6))
    x = np.arange(len(df_summary))
    width = 0.35

    ax.bar(x - width/2, df_summary['Budget_Limit'], width, label='Budget', color='#2E7D32')
    ax.bar(x + width/2, df_summary['Actual_Spend'], width, label='Actual', color='#1565C0')

    ax.set_xlabel('Category', fontsize=11)
    ax.set_ylabel('Amount ($)', fontsize=11)
    ax.set_title('Budget vs Actual Spending by Category', fontsize=13, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(df_summary['Category'], rotation=45, ha='right')
    ax.legend()
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()

    chart1 = os.path.join(output_dir, 'day80_budget_vs_actual.png')
    plt.savefig(chart1, dpi=150)
    plt.close()

    # Chart 2: Variance
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ['#C62828' if v > 0 else '#2E7D32' for v in df_summary['Variance']]
    ax.bar(df_summary['Category'], df_summary['Variance'], color=colors)
    ax.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
    ax.set_xlabel('Category', fontsize=11)
    ax.set_ylabel('Variance ($)', fontsize=11)
    ax.set_title('Budget Variance by Category (Red = Over, Green = Under)', fontsize=13, fontweight='bold')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    chart2 = os.path.join(output_dir, 'day80_variance.png')
    plt.savefig(chart2, dpi=150)
    plt.close()

    logger.info(f"Charts saved to {output_dir}")
    return chart1, chart2


# =============================================================================
# STEP 7: FORMATTED MULTI-SHEET EXCEL EXPORT
# =============================================================================
def export_report(output_path, df_raw, df_processed, df_summary, chart_paths):
    """Export professional multi-sheet Excel report with embedded charts."""
    logger.info(f"Exporting report to {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Raw Data
        df_raw.to_excel(writer, sheet_name='01_Raw_Data', index=False)

        # Sheet 2: Processed Data
        df_processed.to_excel(writer, sheet_name='02_Processed', index=False)

        # Sheet 3: Summary Dashboard
        df_summary.to_excel(writer, sheet_name='03_Summary', index=False)

        # Formatting: Summary Sheet
        ws = writer.sheets['03_Summary']
        header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Color status text
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                if cell.value == 'Over Budget':
                    cell.font = Font(color='C62828', bold=True)
                elif cell.value == 'Under Budget':
                    cell.font = Font(color='2E7D32', bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

        # Embed charts
        try:
            if chart_paths and len(chart_paths) >= 2:
                img1 = OpenpyxlImage(chart_paths[0])
                img1.anchor = 'H2'
                ws.add_image(img1)

                img2 = OpenpyxlImage(chart_paths[1])
                img2.anchor = 'H22'
                ws.add_image(img2)
        except Exception as e:
            logger.warning(f"Chart embedding skipped: {e}")

    logger.info("Report export complete")


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    logger.info("=== Day 80 Build: Expense Tracker ===")

    # Ensure input exists
    if not os.path.exists(INPUT_FILE):
        create_practice_sheet(INPUT_FILE)

    # Pipeline
    df_raw, df_budget = load_data(INPUT_FILE)
    df_processed, df_summary = process_expenses(df_raw, df_budget, CATEGORY_RULES)
    generate_insights(df_summary)
    chart_paths = generate_visualizations(df_summary, CHART_DIR)
    export_report(OUTPUT_FILE, df_raw, df_processed, df_summary, chart_paths)

    # Console summary
    print("\n" + "="*50)
    print("DAY 80 BUILD COMPLETE")
    print("="*50)
    print(f"Transactions Processed : {len(df_processed)}")
    print(f"Categories Assigned    : {df_processed['Category'].nunique()}")
    print(f"Output Report          : {OUTPUT_FILE}")
    print(f"Chart Assets           : {CHART_DIR}/")
    print(f"Execution Log          : expense_tracker_day80.log")
    print("="*50)


if __name__ == '__main__':
    main()